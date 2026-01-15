"""
进度跟踪模块
负责读取任务和记录爬取进度
"""

import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from pathlib import Path
import logging
from typing import List, Dict, Optional
from datetime import datetime

logger = logging.getLogger(__name__)


class ProgressTracker:
    """进度跟踪管理器"""

    def __init__(self, excel_path: str, progress_path: str = "爬取进度.xlsx"):
        """
        初始化进度跟踪器

        Args:
            excel_path: 源Excel文件路径
            progress_path: 进度Excel文件路径
        """
        self.excel_path = Path(excel_path)
        self.progress_path = Path(progress_path)
        self.tasks: List[Dict] = []
        self.progress_data: Dict[str, Dict] = {}

    def load_tasks(self) -> List[Dict]:
        """
        从源Excel加载任务列表

        Returns:
            任务列表，每个任务包含school_name和url字段
        """
        logger.info(f"正在加载任务列表: {self.excel_path}")

        wb = load_workbook(self.excel_path, read_only=True, data_only=True)
        ws = wb.active

        # 查找列索引
        header_row = None
        name_col = None
        url_col = None

        for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
            if any(cell and ("院校名称" in str(cell) or "学校名称" in str(cell)) for cell in row if cell):
                header_row = row_idx
                for col_idx, cell in enumerate(row):
                    if cell and ("院校名称" in str(cell) or "学校名称" in str(cell)):
                        name_col = col_idx
                    elif cell and "招生章程详情页链接（本部）" in str(cell):
                        url_col = col_idx
                break

        if header_row is None or name_col is None or url_col is None:
            raise ValueError("无法找到必需的列：学校名称 或 招生章程详情页链接（本部）")

        # 读取数据
        tasks = []
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            if not row or len(row) <= max(name_col, url_col):
                continue

            school_name = row[name_col]
            url = row[url_col]

            # 跳过空URL
            if not school_name or not url:
                continue

            # 跳过无效URL
            if not isinstance(url, str) or not url.startswith("http"):
                continue

            tasks.append({
                "school_name": str(school_name).strip(),
                "url": str(url).strip()
            })

        wb.close()
        self.tasks = tasks

        logger.info(f"加载完成: 共 {len(tasks)} 个任务")
        return tasks

    def init_progress_file(self):
        """初始化进度Excel文件"""
        if self.progress_path.exists():
            logger.info(f"进度文件已存在: {self.progress_path}")
            self._load_progress()
            return

        # 创建新进度文件
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "爬取进度"

        # 设置表头
        headers = ["序号", "院校名称", "详情页链接", "状态", "字符数", "表格数", "爬取时间", "备注"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # 保存
        wb.save(self.progress_path)
        logger.info(f"已创建进度文件: {self.progress_path}")

    def _load_progress(self):
        """从进度文件加载已记录的进度"""
        wb = load_workbook(self.progress_path, read_only=True, data_only=True)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            school_name = row[1]
            if school_name:
                self.progress_data[school_name] = {
                    "status": row[3],
                    "text_length": row[4],
                    "table_count": row[5],
                    "crawl_time": row[6],
                    "note": row[7]
                }

        wb.close()
        logger.info(f"已加载进度: {len(self.progress_data)} 条记录")

    def update_progress(self, school_name: str, status: str, text_length: int = 0,
                       table_count: int = 0, note: str = ""):
        """
        更新单个任务进度

        Args:
            school_name: 学校名称
            status: 状态（成功/失败/跳过）
            text_length: 文本字符数
            table_count: 表格数量
            note: 备注
        """
        # 更新内存数据
        self.progress_data[school_name] = {
            "status": status,
            "text_length": text_length,
            "table_count": table_count,
            "crawl_time": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "note": note
        }

        # 写入Excel
        self._write_progress_to_excel()

    def _write_progress_to_excel(self):
        """将进度数据写入Excel文件"""
        wb = load_workbook(self.progress_path)
        ws = wb.active

        # 清除现有数据（保留表头）
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.value = None

        # 写入数据
        row_idx = 2
        for idx, task in enumerate(self.tasks, 1):
            school_name = task["school_name"]
            url = task["url"]

            # 获取进度数据
            progress = self.progress_data.get(school_name, {})

            ws.cell(row=row_idx, column=1, value=idx)
            ws.cell(row=row_idx, column=2, value=school_name)
            ws.cell(row=row_idx, column=3, value=url)
            ws.cell(row=row_idx, column=4, value=progress.get("status", "待爬取"))
            ws.cell(row=row_idx, column=5, value=progress.get("text_length", 0))
            ws.cell(row=row_idx, column=6, value=progress.get("table_count", 0))
            ws.cell(row=row_idx, column=7, value=progress.get("crawl_time", ""))
            ws.cell(row=row_idx, column=8, value=progress.get("note", ""))

            # 根据状态设置颜色
            status = progress.get("status", "待爬取")
            if status == "成功":
                fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            elif status == "失败":
                fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            elif status == "跳过":
                fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
            else:
                fill = None

            if fill:
                for col in range(1, 9):
                    ws.cell(row=row_idx, column=col).fill = fill

            row_idx += 1

        # 保存
        wb.save(self.progress_path)

    def get_remaining_tasks(self, storage_dir: str = "details") -> List[Dict]:
        """
        获取剩余未完成的任务

        Args:
            storage_dir: MD文件保存目录

        Returns:
            剩余任务列表
        """
        remaining = []
        storage_path = Path(storage_dir)

        for task in self.tasks:
            school_name = task["school_name"]
            md_file = storage_path / f"{school_name}.md"

            # 检查MD文件是否存在
            if md_file.exists():
                continue

            # 检查进度记录
            progress = self.progress_data.get(school_name, {})
            if progress.get("status") in ["成功", "跳过"]:
                continue

            remaining.append(task)

        logger.info(f"剩余任务数: {len(remaining)} / {len(self.tasks)}")
        return remaining

    def get_statistics(self) -> Dict:
        """获取统计信息"""
        total = len(self.tasks)
        completed = sum(1 for p in self.progress_data.values() if p.get("status") == "成功")
        failed = sum(1 for p in self.progress_data.values() if p.get("status") == "失败")
        pending = total - completed - failed

        return {
            "total": total,
            "completed": completed,
            "failed": failed,
            "pending": pending,
            "success_rate": f"{completed / total * 100:.2f}%" if total > 0 else "0%"
        }
