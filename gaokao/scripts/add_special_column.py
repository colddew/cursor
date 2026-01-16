#!/usr/bin/env python3
"""
在招生章程（详情）.xlsx中新增"招生章程内容（特殊）"列
只填入3所特殊学校的内容，保持原有格式和样式不变
"""

from openpyxl import load_workbook
from pathlib import Path
import pandas as pd

# 文件路径
EXCEL_FILE = "achievement/招生章程（详情）.xlsx"
SPECIAL_DIR = "special"

# 3所特殊学校
SPECIAL_SCHOOLS = [
    "首都经济贸易大学",
    "浙江工商大学",
    "西南林业大学"
]

print("=" * 60)
print("开始添加特殊招生章程列")
print("=" * 60)

# 读取Excel获取学校名称列和行号对应关系
df = pd.read_excel(EXCEL_FILE)
school_column_name = "学校名称"

print(f"\n总学校数: {len(df)}")

# 找到3所特殊学校的行号（从1开始，因为有表头）
special_rows = {}
for idx, row in df.iterrows():
    school_name = row[school_column_name]
    if school_name in SPECIAL_SCHOOLS:
        # Excel行号 = idx + 2（idx从0开始，加1是表头，再加1是数据行）
        special_rows[school_name] = idx + 2
        print(f"  找到: {school_name} (行{special_rows[school_name]})")

print(f"\n找到3所特殊学校")

# 使用openpyxl打开文件，保留格式
print(f"\n打开Excel文件: {EXCEL_FILE}")
wb = load_workbook(EXCEL_FILE)
ws = wb.active

# 获取最大列数（从第1行开始）
max_col = ws.max_column
max_row = ws.max_row

print(f"当前列数: {max_col}")
print(f"当前行数: {max_row}")

# 在最后一列后面新增一列
new_col_num = max_col + 1
new_col_letter = chr(64 + new_col_num) if new_col_num <= 26 else chr(64 + (new_col_num - 1) // 26) + chr(65 + (new_col_num - 1) % 26)

# 设置列标题
ws.cell(row=1, column=new_col_num, value="招生章程内容（特殊）")

print(f"\n新增列: {new_col_letter}列")

# 获取special文件夹中的MD文件
special_files = {
    "首都经济贸易大学": None,
    "浙江工商大学": None,
    "西南林业大学": None
}

# 查找MD文件
for md_file in Path(SPECIAL_DIR).glob("*.md"):
    for school_name in SPECIAL_SCHOOLS:
        if school_name in md_file.name:
            special_files[school_name] = md_file
            print(f"  找到文件: {md_file.name}")
            break

# 填入3所特殊学校的内容
filled_count = 0
for school_name, row_num in special_rows.items():
    md_file = special_files.get(school_name)
    if md_file and md_file.exists():
        with open(md_file, 'r', encoding='utf-8') as f:
            content = f.read()
        ws.cell(row=row_num, column=new_col_num, value=content)
        print(f"  ✓ {school_name}: {len(content)}字符")
        filled_count += 1
    else:
        print(f"  ✗ {school_name}: 文件不存在")

# 保存文件
print(f"\n保存文件: {EXCEL_FILE}")
wb.save(EXCEL_FILE)

print("\n" + "=" * 60)
print(f"完成！新增了1列，填入了{filled_count}所特殊学校的内容")
print("=" * 60)
