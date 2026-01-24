#!/usr/bin/env python3
"""
AI Studio PaddleOCR-VL 调用脚本
使用 .env 文件管理 API 凭证
"""

import sys
# 禁用生成 __pycache__
sys.dont_write_bytecode = True
import base64
import requests
from pathlib import Path
import time
import os
import argparse
from dotenv import load_dotenv

# 自动定位项目根目录并加入 sys.path (支持脚本直接调用)
script_dir = Path(__file__).resolve().parent
project_root = script_dir.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from scripts.common.baidu_api_ocr import call_paddleocr_vl
from scripts.common.excel_utils import html_table_to_excel, BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter

# -------------------------------------------------------------------------
# 加载 .env 文件
load_dotenv()



def save_results(result, output_dir, base_filename):
    """保存识别结果（使用原文件名+时间戳）"""
    # 强制确保目录为 Path 对象并存在
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)
    saved_files = []
    
    # 生成时间戳
    timestamp = time.strftime("%Y%m%d_%H%M%S")
    
    for i, res in enumerate(result.get("layoutParsingResults", [])):
        markdown_obj = res.get("markdown", {})
        
        # 保存 Markdown（原文件名_时间戳.md）
        md_text = markdown_obj.get("text", "")
        if md_text:
            md_filename = os.path.join(output_dir, f"{base_filename}_{timestamp}.md")
            with open(md_filename, "w", encoding="utf-8") as md_file:
                md_file.write(md_text)
            saved_files.append((md_filename, md_text))
            print(f"   ✓ Markdown: {md_filename}")
        
        # 保存图片
        for img_path, img_url in markdown_obj.get("images", {}).items():
            full_img_path = os.path.join(output_dir, img_path)
            os.makedirs(os.path.dirname(full_img_path), exist_ok=True)
            img_bytes = requests.get(img_url).content
            with open(full_img_path, "wb") as img_file:
                img_file.write(img_bytes)
            # named_img_filename = f"{base_filename}_{timestamp}.jpg"
            print(f"   ✓ 图片: {full_img_path}")
    
    return saved_files, timestamp

def markdown_to_excel(markdown_text, output_path):
    """
    将 Markdown/HTML 中的表格解析为 Excel (浙江专用 - 已重构为通用逻辑)
    """
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter

        soup = BeautifulSoup(markdown_text, 'html.parser')
        if not soup.find('table'):
            return False
            
        wb = Workbook()
        ws = wb.active
        ws.title = "识别结果"
        
        # 调用共享的解析逻辑
        _, max_cols = html_table_to_excel(soup, ws)
        
        # 自适应列宽逻辑保持在脚本中，因为可能需要省份特定的宽度控制
        max_col_width = int(os.getenv('EXCEL_MAX_COLUMN_WIDTH', '60'))
        for col_idx in range(1, max_cols + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        cell_str = str(cell.value)
                        display_width = sum(2 if ord(c) > 127 else 1 for c in cell_str)
                        max_length = max(max_length, display_width)
            if max_length > 0:
                adjusted_width = min(max_length + 2, max_col_width)
                ws.column_dimensions[col_letter].width = adjusted_width
                
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"Excel conversion failed: {e}")
        return False

def main():
    parser = argparse.ArgumentParser(description="Call AI Studio PaddleOCR-VL API")
    parser.add_argument("image_path", help="Path to the image file")
    parser.add_argument("--output-dir", help="Directory to save output files", default=None)
    parser.add_argument("--province", help="Province code for strategy selection", default="zhejiang")
    parser.add_argument("--model", choices=['auto', 'ocr', 'v3'], default='ocr',
                        help="OCR 模型选择 (浙江目前仅支持 ocr 逻辑)")
    
    args = parser.parse_args()
    
    file_path = Path(args.image_path)
    if not file_path.exists():
        print(f"Error: File not found {file_path}")
        sys.exit(1)
        
    # 处理输出目录
    if args.output_dir:
        output_dir = Path(args.output_dir)
    else:
        # 默认：project_root/output
        output_dir = Path(os.getcwd()) / "output"
    
    output_dir.mkdir(parents=True, exist_ok=True)
    base_filename = file_path.stem

    # 获取 API 配置
    api_url = os.getenv("AISTUDIO_API_URL")
    token = os.getenv("AISTUDIO_TOKEN")
    
    if not api_url or not token:
        print("❌ Error: AISTUDIO_API_URL or AISTUDIO_TOKEN not found in .env file")
        sys.exit(1)
        
    print(f"=" * 60)
    print(f"📄 输入: {file_path.name}")
    print(f"🌍 策略: {args.province}")
    print(f"🔗 API: {api_url[:50]}...")
    print(f"📁 输出: {output_dir}/")
    
    total_start = time.time()
    
    print(f"\n🔍 调用 API...", end='', flush=True)
    
    try:
        response = call_paddleocr_vl(str(file_path), api_url, token)
        
        if response.status_code != 200:
            print(f" 失败")
            print(f"   ❌ HTTP {response.status_code}: {response.text}")
            sys.exit(1)
        
        result_json = response.json()
        
        if result_json.get("errorCode", 0) != 0:
            print(f" 失败")
            print(f"   ❌ {result_json.get('errorMsg', 'Unknown error')}")
            sys.exit(1)
        
        ocr_time = time.time() - total_start
        print(f" 成功（{ocr_time:.1f}秒）")
        
        result = result_json.get("result", {})
        
        print(f"\n📝 保存结果:")
        saved_files, timestamp = save_results(result, output_dir, base_filename)
        
        # 转换为 Excel（单文件模式）
        excel_saved = False
        for file_path_str, md_text in saved_files:
            if md_text and file_path_str.endswith('.md'):
                excel_path = Path(output_dir) / f"{base_filename}_{timestamp}.xlsx"
                if markdown_to_excel(md_text, excel_path):
                    print(f"   ✓ Excel: {excel_path}")
                    excel_saved = True
                else:
                    print(f"   ⚠️ Excel 生成失败 (浙江专用代码)")
        
        total_time = time.time() - total_start
        
        print(f"\n🎉 完成！")
        print(f"⏱️  总耗时: {total_time:.1f}秒")
        print(f"📊 共保存 {len(saved_files) + (1 if excel_saved else 0)} 个文件")
        print(f"📁 输出目录: {Path(output_dir).absolute()}")
        
    except Exception as e:
        print(f" 失败")
        print(f"   ❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

if __name__ == "__main__":
    main()
