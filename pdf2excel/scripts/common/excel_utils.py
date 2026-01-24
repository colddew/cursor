import os
import sys
# 禁用生成 __pycache__
sys.dont_write_bytecode = True
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup

def get_header_style():
    """获取表头样式"""
    bg_color = os.getenv('EXCEL_HEADER_BG_COLOR', 'CCE5FF')
    header_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
    header_font = Font(bold=True)
    return header_fill, header_font

def get_common_styles():
    """获取通用对齐和边框样式"""
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    return align_center, align_left, border

def setup_columns(ws):
    """配置 Excel 列宽 (读取环境变量，若无则使用默认值)"""
    width_a = float(os.getenv('EXCEL_COL_WIDTH_A', '20'))
    width_b = float(os.getenv('EXCEL_COL_WIDTH_B', '40'))
    width_c = float(os.getenv('EXCEL_COL_WIDTH_C', '60'))
    
    ws.column_dimensions['A'].width = width_a
    ws.column_dimensions['B'].width = width_b
    ws.column_dimensions['C'].width = width_c
    # print(f"   📏 已应用固定列宽: A={width_a}, B={width_b}, C={width_c}")

def write_data_row(ws, row_idx, data, styles):
    """写入一行数据并应用样式"""
    align_center, align_left, border = styles
    for col, val in enumerate(data, 1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.alignment = align_center if col == 1 else align_left
        cell.border = border

def write_text_row(ws, row_idx, text, alignment):
    """写入一行合并单元格的文本"""
    ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=3)
    cell = ws.cell(row=row_idx, column=1, value=text)
    cell.alignment = alignment

def html_table_to_excel(soup, ws, start_row=1):
    """
    将 HTML <table> 标签解析并写入 Excel 工作表 (通用逻辑)
    返回下一行起始行号和最大列数
    """
    html_tables = soup.find_all('table')
    if not html_tables:
        return start_row, 1

    # 配置信息
    gap_rows = int(os.getenv('EXCEL_TABLE_GAP_ROWS', '3'))
    max_col_width = int(os.getenv('EXCEL_MAX_COLUMN_WIDTH', '60'))

    # 获取最大列数
    max_cols = 1
    for table in html_tables:
        for tr in table.find_all('tr'):
            col_count = 0
            for cell in tr.find_all(['td', 'th']):
                col_count += int(cell.get('colspan', 1))
            max_cols = max(max_cols, col_count)

    occupied = set()
    current_row = start_row
    last_was_table = False

    for element in soup.contents:
        if not element or (hasattr(element, 'name') and element.name == None):
            continue
        
        if getattr(element, 'name', None) == 'table':
            if last_was_table:
                current_row += gap_rows
            
            last_was_table = True
            rows = element.find_all('tr')
            table_start_row = current_row
            
            for r_idx_in_table, tr in enumerate(rows, 1):
                r_idx = table_start_row + r_idx_in_table - 1
                c_idx = 1
                cells = tr.find_all(['td', 'th'])
                
                for cell in cells:
                    while (r_idx, c_idx) in occupied:
                        c_idx += 1
                    
                    rowspan = int(cell.get('rowspan', 1))
                    colspan = int(cell.get('colspan', 1))
                    cell_value = cell.get_text(strip=True)
                    
                    excel_cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                    
                    # 样式应用
                    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
                    is_header = (cell.name == 'th' or r_idx_in_table == 1)
                    alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
                    font = Font(bold=True) if is_header else None
                    
                    for r in range(r_idx, r_idx + rowspan):
                        for c in range(c_idx, c_idx + colspan):
                            target_cell = ws.cell(row=r, column=c)
                            target_cell.border = border
                            target_cell.alignment = alignment
                            if font: target_cell.font = font
                            occupied.add((r, c))
                    
                    if rowspan > 1 or colspan > 1:
                        ws.merge_cells(start_row=r_idx, start_column=c_idx,
                                     end_row=r_idx + rowspan - 1, end_column=c_idx + colspan - 1)
                    c_idx += colspan
            
            current_row = max([r for r, c in occupied] if occupied else [current_row]) + 1
            
        else:
            # 处理非表格文本 (可能是标题或备注)
            text_content = element.get_text(strip=True) if hasattr(element, 'get_text') else str(element).strip()
            if text_content and len(text_content) < int(os.getenv('MAX_TEXT_LABEL_LENGTH', '100')):
                last_was_table = False
                current_row += 1
                ws.cell(row=current_row, column=1, value=text_content)
                if max_cols > 1:
                    ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
                
                title_cell = ws.cell(row=current_row, column=1)
                title_cell.font = Font(bold=True)
                title_cell.alignment = Alignment(horizontal='center', vertical='center')
                
                for c in range(1, max_cols + 1):
                    occupied.add((current_row, c))
                current_row += 1

    return current_row, max_cols

def autofit_columns(ws, max_cols, max_width=60):
    """
    自动调整 Excel 列宽 (浙江/安徽通用)
    原理：遍历每一列，计算最长内容的显示宽度(中文字符x2)，动态设置列宽。
    """
    for col_idx in range(1, max_cols + 1):
        col_letter = get_column_letter(col_idx)
        max_length = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value:
                    cell_str = str(cell.value)
                    # 计算显示宽度（中文字符计2，英文计1）
                    display_width = sum(2 if ord(c) > 127 else 1 for c in cell_str)
                    max_length = max(max_length, display_width)
        if max_length > 0:
            # 留一点余量 (+2)，但不超过最大限制
            adjusted_width = min(max_length + 2, max_width)
            ws.column_dimensions[col_letter].width = adjusted_width
