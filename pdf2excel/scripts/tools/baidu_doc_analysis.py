#!/usr/bin/env python3
import sys; sys.dont_write_bytecode = True"""
百度文档解析服务 - 2026年最新
支持表格识别、版面分析、阅读顺序等
服务位置：百度AI开放平台 > 文字识别 > 文档解析
"""

import sys
import base64
import requests
from pathlib import Path
import time
import json
from dotenv import load_dotenv

# 加载环境变量
load_dotenv()

def get_baidu_token(api_key, secret_key):
    """获取百度 access_token"""
    url = "https://aip.baidubce.com/oauth/2.0/token"
    params = {
        "grant_type": "client_credentials",
        "client_id": api_key,
        "client_secret": secret_key
    }
    
    response = requests.post(url, params=params)
    if response.status_code == 200:
        return response.json().get("access_token")
    else:
        raise Exception(f"获取 token 失败: {response.text}")

def doc_analysis(image_path, access_token):
    """
    使用百度文档解析服务
    API: https://cloud.baidu.com/doc/OCR/s/1l3h7y4ky
    """
    url = f"https://aip.baidubce.com/rest/2.0/solution/v1/doc_analysis/doc_analysis?access_token={access_token}"
    
    # 读取图片并 base64 编码
    with open(image_path, 'rb') as f:
        image_data = base64.b64encode(f.read()).decode('utf-8')
    
    headers = {'Content-Type': 'application/x-www-form-urlencoded'}
    data = {
        'image': image_data,
        'doc_analysis_type': 'table',  # 专门针对表格
        'return_format': 'markdown'  # 返回 Markdown 格式（更易处理）
    }
    
    response = requests.post(url, headers=headers, data=data)
    
    if response.status_code == 200:
        result = response.json()
        if 'results' in result:
            return result
        else:
            raise Exception(f"识别失败: {result.get('error_msg', result)}")
    else:
        raise Exception(f"API 调用失败: {response.text}")

def markdown_to_excel(markdown_text, output_path):
    """将 Markdown/HTML 表格转换为 Excel"""
    from openpyxl import Workbook
    from openpyxl.styles import Border, Side, Alignment, Font
    from bs4 import BeautifulSoup
    
    # 先尝试解析 HTML 表格
    soup = BeautifulSoup(markdown_text, 'html.parser')
    html_tables = soup.find_all('table')
    
    if html_tables:
        wb = Workbook()
        wb.remove(wb.active)
        
        for idx, table in enumerate(html_tables):
            ws = wb.create_sheet(f"表格{idx + 1}")
            occupied = set()
            rows = table.find_all('tr')
            for r_idx, tr in enumerate(rows, 1):
                c_idx = 1
                cells = tr.find_all(['td', 'th'])
                for cell in cells:
                    while (r_idx, c_idx) in occupied:
                        c_idx += 1
                    
                    rowspan = int(cell.get('rowspan', 1))
                    colspan = int(cell.get('colspan', 1))
                    cell_value = cell.get_text(strip=True)
                    excel_cell = ws.cell(row=r_idx, column=c_idx, value=cell_value)
                    
                    border = Border(
                        left=Side(style='thin'),
                        right=Side(style='thin'),
                        top=Side(style='thin'),
                        bottom=Side(style='thin')
                    )
                    alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')
                    font = None
                    if cell.name == 'th' or r_idx == 1:
                        font = Font(bold=True)
                    
                    for r in range(r_idx, r_idx + rowspan):
                        for c in range(c_idx, c_idx + colspan):
                            target_cell = ws.cell(row=r, column=c)
                            target_cell.border = border
                            target_cell.alignment = alignment
                            if font:
                                target_cell.font = font
                            occupied.add((r, c))
                    
                    if rowspan > 1 or colspan > 1:
                        ws.merge_cells(
                            start_row=r_idx, 
                            start_column=c_idx, 
                            end_row=r_idx + rowspan - 1, 
                            end_column=c_idx + colspan - 1
                        )
                    c_idx += colspan
        
        wb.save(output_path)
        return True
    
    # 如果没有 HTML 表格，尝试 Markdown 表格
    tables = []
    current_table = []
    in_table = False
    
    for line in markdown_text.split('\n'):
        line = line.strip()
        if line.startswith('|'):
            in_table = True
            if not line.replace('|', '').replace('-', '').replace(':', '').strip():
                continue
            current_table.append(line)
        else:
            if in_table and current_table:
                tables.append(current_table)
                current_table = []
                in_table = False
    
    if current_table:
        tables.append(current_table)
    
    if not tables:
        return False
    
    wb = Workbook()
    wb.remove(wb.active)
    
    for idx, table_lines in enumerate(tables):
        ws = wb.create_sheet(f"表格{idx + 1}")
        for row_idx, line in enumerate(table_lines, 1):
            cells = [cell.strip() for cell in line.split('|')[1:-1]]
            for col_idx, cell_value in enumerate(cells, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=cell_value)
                cell.border = Border(
                    left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin')
                )
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                if row_idx == 1:
                    cell.font = Font(bold=True)
    
    wb.save(output_path)
    return True

def extract_with_doc_analysis():
    """使用百度文档解析服务"""
    import os
    
    api_key = os.getenv('BAIDU_OCR_API_KEY')
    secret_key = os.getenv('BAIDU_OCR_SECRET_KEY')
    
    if not api_key or not secret_key:
        print("\n❌ 请设置百度 OCR API Key:")
        print("   export BAIDU_OCR_API_KEY='your_api_key'")
        print("   export BAIDU_OCR_SECRET_KEY='your_secret_key'")
        print("\n💡 获取方式:")
        print("   1. 访问 https://ai.baidu.com/")
        print("   2. 登录并完成实名认证")
        print("   3. 创建应用，选择【文档解析】服务")
        print("   4. 获取 API Key 和 Secret Key")
        print("\n📊 免费额度（2026年）:")
        print("   - 个人认证: 500次/月")
        print("   - 企业认证: 1000次/月")
        sys.exit(1)
    
    if len(sys.argv) < 2:
        print("用法: python baidu_doc_analysis.py <图片文件>")
        sys.exit(1)
    
    input_path = Path(sys.argv[1])
    
    # 统一输出到项目根目录的 output
    project_root = Path(__file__).parent.parent
    output_dir = project_root / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    
    if len(sys.argv) > 2:
        output_path = Path(sys.argv[2])
    else:
        output_path = output_dir / input_path.with_suffix('.xlsx').name
    
    if not input_path.exists():
        print(f"❌ 文件不存在: {input_path}")
        sys.exit(1)
    
    print(f"\n🚀 百度文档解析服务 (2026版)")
    print(f"=" * 60)
    print(f"📄 输入: {input_path.name}")
    print(f"📊 输出: {output_path}")
    
    total_start = time.time()
    
    # 获取 access_token
    print("\n🔑 获取访问令牌...")
    try:
        access_token = get_baidu_token(api_key, secret_key)
        print("   ✓ 成功")
    except Exception as e:
        print(f"   ❌ {e}")
        sys.exit(1)
    
    # 文档解析
    print(f"\n📄 解析文档...")
    print(f"   🔍 处理中...", end='', flush=True)
    
    try:
        result = doc_analysis(str(input_path), access_token)
        ocr_time = time.time() - total_start
        
        print(f" 完成（{ocr_time:.1f}秒）")
        
        # 提取结果
        results = result.get('results', [])
        
        if results:
            # 合并所有结果（可能包含文字和表格）
            all_content = []
            table_count = 0
            
            for item in results:
                result_type = item.get('type', '')
                content = item.get('content', '')
                
                if result_type == 'table':
                    table_count += 1
                    all_content.append(content)
                else:
                    all_content.append(content)
            
            markdown_text = '\n\n'.join(all_content)
            
            print(f"   ✓ 识别到 {table_count} 个表格区域")
            
            # 保存 Markdown（方便查看）
            md_path = output_path.with_suffix('.md')
            with open(md_path, 'w', encoding='utf-8') as f:
                f.write(markdown_text)
            print(f"   ✓ Markdown 已保存: {md_path.name}")
            
            # 转换为 Excel
            if markdown_to_excel(markdown_text, output_path):
                print(f"\n🎉 成功！文件: {output_path.absolute()}")
                print(f"⏱️  总耗时: {ocr_time:.1f}秒")
                print(f"📊 包含边框和格式")
            else:
                print(f"\n⚠️  未检测到表格，已保存为 Markdown")
        else:
            print(f"\n⚠️  未识别到内容")
        
    except Exception as e:
        print(f" 失败")
        print(f"   ❌ {e}")
        sys.exit(1)

if __name__ == "__main__":
    try:
        extract_with_doc_analysis()
    except KeyboardInterrupt:
        print("\n\n⚠️  用户中断")
        sys.exit(1)
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)
