#!/usr/bin/env python3
"""
安徽省物理组/招生大本专用处理脚本 (基于 StructureV3)
1. 调用 StructureV3 接口获取阅读顺序正确的文本
2. 使用正则表达式解析三列数据 ([代码] [名称] [人数/备注])
3. 处理跨行备注及多行专业详细说明
4. 过滤说明性文字及优化行间距
"""

import sys
# 禁用生成 __pycache__
sys.dont_write_bytecode = True
import os
import re
import time
import argparse
from pathlib import Path
from dotenv import load_dotenv

# 加载配置 (尽量在其他导入前执行)
load_dotenv()

from openpyxl import Workbook
from openpyxl.styles import Font

# 自动定位项目根目录并加入 sys.path
script_dir = Path(__file__).resolve().parent
project_root = script_dir.parent.parent
if str(project_root) not in sys.path:
    sys.path.insert(0, str(project_root))

from scripts.common.baidu_api import call_structure_v3
from scripts.common.baidu_api_ocr import call_paddleocr_vl
from scripts.common.excel_utils import (
    get_header_style, get_common_styles, setup_columns, 
    write_data_row, write_text_row, html_table_to_excel,
    autofit_columns, BeautifulSoup
)
from scripts.common.table_detection import detect_table

class AnhuiProcessor:
    # 正则规则 (兼容 StructureV3 的紧凑模式，移除行首锚点以支持非对齐文本)
    GROUP_PATTERN = re.compile(r"(\d{3,})\s*(.*?)\s*(\d+\s*人.*)")
    
    # Code组变为可选，名称组排除数字开头
    MAJOR_PATTERN = re.compile(r"([A-Za-z0-9]{2})?\s*([^\d\n\s]{2,}[^\s人]*)\s*(\d+\s*人.*)")
    
    # 填报说明特征模式 (基于结构描述特征)
    INSTRUCTION_FEATURES = [
        r"前\s*\d+\s*位数字为.*代码",
        r"含\s*\d+\s*位专业组代码",
        r"括号内为专业收费标准",
        r"专业名称后数字为.*人数",
        r"院校名称后为专业组"
    ]

    def __init__(self):
        self.api_url = os.getenv('AISTUDIO_STRUCTURE_URL')
        self.token = os.getenv('AISTUDIO_STRUCTURE_TOKEN') or os.getenv('AISTUDIO_TOKEN')

    def is_instruction_line(self, line):
        """判断是否为填报说明类型的无效文字"""
        # 优先判断：如果是数据行（以数字代码开头且包含人数），则不是说明文字
        if re.match(r'^\d{2,4}\s+', line) and re.search(r'\d+\s*人', line):
            return False
        
        # 检查特征模式
        for feature in self.INSTRUCTION_FEATURES:
            if re.search(feature, line):
                return True
        
        # 额外特征词检查（Page 05 特有的说明文字）
        instruction_keywords = [
            "考生本人务必", "填报志愿", "志愿信息", "录入错误",
            "招生计划不分", "科目组合", "综合分计算公式", "综合分=",
            "平行志愿", "院校专业组志愿", "专业服从志愿",
            "$$"  # 数学公式标记
        ]
        if any(kw in line for kw in instruction_keywords):
            return True
        
        # 超长段落检查：超过100字符很可能是说明文字（不再要求不包含"人"）
        if len(line) > 100:
            return True
        
        return False

    def run(self, image_path, output_dir):
        if not self.api_url or not self.token:
            print("❌ 错误: 请在 .env 中配置 AISTUDIO_STRUCTURE_URL 和 AISTUDIO_STRUCTURE_TOKEN")
            return

        print(f"🔍 正在处理 (Anhui): {Path(image_path).name}")
        
        # 1. OpenCV 预检测：表格还是文本？
        # 根据 model_preference 决定是否强制使用某个模型
        force_mode = getattr(self, 'model_preference', 'auto')
        if force_mode == 'auto':
            force_mode = None  # 自动检测
        
        is_table, intersections = detect_table(image_path, force_mode=force_mode)
        
        if force_mode:
            mode_str = f"强制模式: {'PaddleOCR' if force_mode == 'ocr' else 'StructureV3'}"
        else:
            mode_str = "表格模式 (PaddleOCR)" if is_table else "文本模式 (StructureV3)"
        print(f"   📊 检测结果: 交点数={intersections} => {mode_str}")
        
        full_text = ""
        
        if is_table:
            # === 表格模式：调用标准 OCR ===
            ocr_api_url = os.getenv("AISTUDIO_API_URL")
            if not ocr_api_url:
                print("❌ 错误: 未在 .env 中配置 AISTUDIO_API_URL (用于表格模式)")
                return
             
            response = call_paddleocr_vl(image_path, ocr_api_url, self.token)
            if response.status_code != 200:
                print(f"❌ OCR API 调用失败: {response.text}")
                return

            res_json = response.json()
            if res_json.get("errorCode", 0) != 0:
                print(f"❌ OCR 业务错误: {res_json.get('errorMsg')}")
                return
            
            layout_results = res_json.get("result", {}).get("layoutParsingResults", [])
            for res in layout_results:
                full_text += res.get("markdown", {}).get("text", "") + "\n"
        else:
            # === 文本模式：调用 StructureV3 (Low Threshold) ===
            response = call_structure_v3(image_path, self.api_url, self.token)
            if response.status_code != 200:
                print(f"❌ V3 API 调用失败: {response.text}")
                return

            res_json = response.json()
            if res_json.get("errorCode", 0) != 0:
                print(f"❌ V3 业务错误: {res_json.get('errorMsg')}")
                return

            layout_results = res_json.get("result", {}).get("layoutParsingResults", [])
            for res in layout_results:
                full_text += res.get("markdown", {}).get("text", "") + "\n"

        # 2. 保存中间结果
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        base_name = Path(image_path).stem
        md_path = Path(output_dir) / f"{base_name}_{timestamp}.md"
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write(full_text)
        print(f"   ✓ Markdown 已保存: {md_path}")

        # 3. 解析为 Excel
        wb = self.parse_to_excel(full_text)
        excel_path = Path(output_dir) / f"{base_name}_{timestamp}.xlsx"
        wb.save(excel_path)
        print(f"   ✓ Excel 已保存: {excel_path}")

    # 定义行类型常量
    LINE_TYPE_TITLE = 'TITLE'
    LINE_TYPE_GROUP = 'GROUP'
    LINE_TYPE_MAJOR = 'MAJOR'
    LINE_TYPE_INSTRUCTION = 'INSTRUCTION'
    LINE_TYPE_UNKNOWN = 'UNKNOWN'

    def classify_line(self, line):
        """第一遍扫描：给每一行打标签"""
        line = line.strip()
        if not line: return None, None
        
        # 1. 标题 - 优先检查（必须在说明文字之前）
        # 特征：以 # 开头（Markdown 标题），或者短且包含"批次"关键词
        if line.startswith('#'):
            # 去掉 # 号，返回干净的标题文本
            clean_title = line.lstrip('#').strip()
            return self.LINE_TYPE_TITLE, clean_title
        if len(line) < 20 and re.search(r'(提前|本科|专科|高职|艺术|体育)批', line):
            return self.LINE_TYPE_TITLE, line
        
        # 2. 明确的说明文字（长段落说明）
        if self.is_instruction_line(line):
            return self.LINE_TYPE_INSTRUCTION, line
            
        # 3. 院校专业组
        # 特征：有3位以上数字代码，且包含"人"（StructureV3特征）
        # GROUP_PATTERN: (\d{3,})\s*(.*?)\s*(\d+\s*人.*)
        if self.GROUP_PATTERN.search(line):
            # 误判防护：如果包含"元/年"等价格特征，这多半是上一行断行的学费信息，不是院校组
            # 兼容半角/和全角／
            if re.search(r'元\s*[\\\/／]\s*年', line):
                return self.LINE_TYPE_UNKNOWN, line
            return self.LINE_TYPE_GROUP, line
            
        # 4. 专业
        # 特征：可选代码 + 名称 + 人数
        # MAJOR_PATTERN: ([A-Za-z0-9]{2})?\s*([^\d\n\s]{2,}[^\s人]*)\s*(\d+\s*人.*)
        # 必须严格验证"人"字存在，避免误判普通文本
        if self.MAJOR_PATTERN.search(line):
            # 同样的误判防护
            if re.search(r'元\s*[\\\/／]\s*年', line):
                return self.LINE_TYPE_UNKNOWN, line
            return self.LINE_TYPE_MAJOR, line
            
        # 5. 未知/其他 (可能是表头乱码，也可能是跨行备注)
        return self.LINE_TYPE_UNKNOWN, line

    def _parse_paddle_format(self, text, wb, ws):
        """解析 PaddleOCR 的简单格式（每行独立）"""
        header_fill, header_font = get_header_style()
        styles = get_common_styles()
        align_center, align_left, border = styles
        
        # === 预处理：合并 OCR 断行 + 过滤说明文字 ===
        lines = text.split('\n')
        merged_lines = []
        i = 0
        while i < len(lines):
            line = lines[i].strip()
            if not line:
                i += 1
                continue
            
            # 【提前过滤】跳过说明文字和特定关键词（避免进入merged_lines）
            if self.is_instruction_line(line):
                i += 1
                continue
            
            if any(kw in line for kw in ["院校名称前", "专业名称前", "招生人数后括号内"]):
                i += 1
                continue
            
            # 检查后续行是否需要合并（跳过空行）
            next_idx = i + 1
            while next_idx < len(lines) and not lines[next_idx].strip():
                next_idx += 1
            
            if next_idx < len(lines):
                next_line = lines[next_idx].strip()
                should_merge = False
                
                # 条件1：下一行以数字+人开头，且当前行不包含"数字+人"模式
                if next_line and re.match(r'^\d+\s*人', next_line) and not re.search(r'\d+\s*人', line):
                    should_merge = True
                
                # 条件2：下一行以括号开头且包含"数字+人"，当前行不包含"数字+人"模式
                # 例如：(思想政治) 2 人
                if next_line and re.match(r'^\(', next_line) and re.search(r'\d+\s*人', next_line) and not re.search(r'\d+\s*人', line):
                    should_merge = True
                
                # 条件3：院校行 + 专业组行合并（仅当院校行没有人数时）
                # 当前行：4位数字开头，无"人"（院校代码+名称，无人数）
                # 下一行：2-3位数字开头，有"人"（专业组代码+人数）
                # 注意：如果当前行已经有"人"，说明是完整的院校专业组行，不应该合并
                current_is_school_no_count = re.match(r'^\d{4}\s+', line) and not re.search(r'\d+\s*人', line)
                next_is_group = re.match(r'^\d{2,3}\s+', next_line) and re.search(r'\d+\s*人', next_line)
                if current_is_school_no_count and next_is_group:
                    should_merge = True
                
                if should_merge:
                    merged_lines.append(line + ' ' + next_line)
                    i = next_idx + 1  # 跳过所有已处理的行
                    continue
            
            merged_lines.append(line)
            i += 1
        
        # === 开始解析 ===
        current_row = 1
        headers_inserted = False
        last_was_title = False  # 跟踪上一行是否是标题
        
        for line in merged_lines:
            line = line.strip()
            if not line:
                continue
            
            # 跳过标题行（以 # 开头或包含"批次"关键词）
            if line.startswith('#'):
                # 标题单独一行，上下加空行
                if current_row > 1:
                    current_row += 1  # 上方空行
                clean_title = line.lstrip('#').strip()
                write_text_row(ws, current_row, clean_title, align_left)
                current_row += 1
                current_row += 1  # 下方空行
                last_was_title = True
                continue
            
            # 跳过说明文字（超长段落 + 特定关键词）
            if self.is_instruction_line(line):
                continue
            
            # 额外过滤：包含"院校名称前"、"专业名称前"等说明性关键词
            if any(kw in line for kw in ["院校名称前", "专业名称前", "招生人数后括号内"]):
                continue
            
            # 只在以下情况插入表头：
            # 1. 第一次遇到数据（headers_inserted=False）
            # 2. 刚遇到过标题（last_was_title=True）
            if not headers_inserted or last_was_title:
                headers = ["专业组/专业代码", "专业组/专业描述", "招生人数/备注"]
                for col, val in enumerate(headers, 1):
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = align_center
                    cell.border = border
                current_row += 1
                headers_inserted = True
                last_was_title = False
            
            # 匹配格式: 数字 + 名称 + X人...
            # 允许名称和人数之间没有空格（如：西班牙语1人）
            match = re.match(r'^(\d{2,4})\s+(.+?)(\d+\s*人.*)$', line)
            
            if match:
                code = match.group(1)
                name = match.group(2).strip()
                remark = match.group(3).strip()
                
                write_data_row(ws, current_row, [code, name, remark], styles)
                current_row += 1
            else:
                # 改进的跨行判断：
                # 情况1：以2-4位数字开头（可能是新数据行的代码），则视为独立数据
                if re.match(r'^\d{2,4}\s+', line):
                    # 这看起来像是一个新的数据行（有代码），尝试解析
                    # 格式可能是：01 专业名称(中外合作办学)(中意合作办学)
                    parts = re.match(r'^(\d{2,4})\s+(.+)$', line)
                    if parts:
                        code = parts.group(1)
                        name = parts.group(2).strip()
                        # 没有人数信息，备注为空
                        write_data_row(ws, current_row, [code, name, ""], styles)
                        current_row += 1
                    else:
                        # 解析失败，追加到上一行
                        if current_row > 2:
                            prev_remark = ws.cell(row=current_row-1, column=3).value or ""
                            ws.cell(row=current_row-1, column=3).value = prev_remark + ' ' + line
                
                # 情况2：包含"X人"模式（可能是跨页数据的延续部分）
                elif re.search(r'\d+\s*人', line):
                    # 这是跨页数据！保留为独立行
                    # 代码为空，整行作为名称+备注
                    # 尝试分离名称和人数
                    parts = re.match(r'^(.+?)(\d+\s*人.*)$', line)
                    if parts:
                        name = parts.group(1).strip()
                        remark = parts.group(2).strip()
                        write_data_row(ws, current_row, ["", name, remark], styles)
                        current_row += 1
                    else:
                        # 无法分离，整行作为备注
                        write_data_row(ws, current_row, ["", "", line], styles)
                        current_row += 1
                
                # 情况3：普通文本，追加到上一行备注
                else:
                    if current_row > 2:
                        prev_remark = ws.cell(row=current_row-1, column=3).value or ""
                        ws.cell(row=current_row-1, column=3).value = prev_remark + ' ' + line
        
        autofit_columns(ws, 3)
        setup_columns(ws)
        return wb

    def parse_to_excel(self, text):
        wb = Workbook()
        ws = wb.active
        
        # 1. 优先尝试 HTML 表格解析
        soup = BeautifulSoup(text, 'html.parser')
        if soup.find('table'):
            print("   📊 检测到 HTML 表格，执行结构化解析...")
            _, max_cols = html_table_to_excel(soup, ws)
            autofit_columns(ws, max_cols)
            if max_cols == 3:
                setup_columns(ws)
            return wb

        # 2. 检测是否是 PaddleOCR 的简单格式（每行独立，格式规整）
        lines = [l.strip() for l in text.split('\n') if l.strip()]
        paddle_format_count = sum(1 for l in lines if re.match(r'^\d{2,4}\s+.+\s+\d+\s*人', l))
        
        if paddle_format_count > len(lines) * 0.5:  # 超过50%的行符合格式
            print("   📝 检测到 PaddleOCR 简单格式，使用行解析...")
            return self._parse_paddle_format(text, wb, ws)

        # 3. 文本模式：上下文感知解析 (2-Pass Strategy)
        header_fill, header_font = get_header_style()
        styles = get_common_styles()
        align_center, align_left, border = styles

        # --- Pass 1: 分类 ---
        raw_lines = text.split('\n')
        tagged_lines = []
        for line in raw_lines:
            l_type, l_content = self.classify_line(line)
            if l_type:
                tagged_lines.append({'type': l_type, 'content': l_content})

        # --- Pass 2: 处理 (状态机) ---
        current_row = 1
        headers_inserted = False
        last_data_row_idx = None # 指向上一条有效数据(Group或Major)，用于追加备注

        # 待写入的非结构化文本缓存
        pending_text_lines = [] 

        def flush_pending_text():
            nonlocal current_row
            if pending_text_lines:
                if current_row > 1: current_row += 1
                for txt in pending_text_lines:
                    write_text_row(ws, current_row, txt, align_left)
                    current_row += 1
                pending_text_lines.clear()

        for i, item in enumerate(tagged_lines):
            l_type = item['type']
            content = item['content']
            
            #以此当前行为中心，查看上下文
            prev_type = tagged_lines[i-1]['type'] if i > 0 else None
            next_type = tagged_lines[i+1]['type'] if i < len(tagged_lines) - 1 else None
            
            # --- 逻辑核心：去噪与归并 ---

            # A. 标题处理
            if l_type == self.LINE_TYPE_TITLE:
                # 特殊去噪：如果当前是"弱标题"(无#)，且下一行也是标题，则认为当前行是页眉/重复标题 -> 丢弃
                if not content.startswith('#') and next_type == self.LINE_TYPE_TITLE:
                    print(f"    🗑️ 丢弃冗余标题: {content}")
                    continue
                
                flush_pending_text()
                # 标题出现，重置数据上下文（不能跨标题合并备注）
                last_data_row_idx = None
                headers_inserted = False # 新段落可能需要新表头
                
                if current_row > 1: current_row += 1
                write_text_row(ws, current_row, content, align_left)
                current_row += 1
                continue

            # B. 说明文字处理
            if l_type == self.LINE_TYPE_INSTRUCTION:
                # 直接忽略
                continue

            # C. 未知行 (UNKNOWN) 处理 - 最关键的部分
            next_type = tagged_lines[i+1]['type'] if i < len(tagged_lines) - 1 else None

            if l_type == self.LINE_TYPE_UNKNOWN:
                # 规则1a: 紧跟在标题前面的 UNKNOWN -> 视为页眉/页脚噪音 (如 "·物理科目...") -> 丢弃
                if next_type == self.LINE_TYPE_TITLE:
                    print(f"    🗑️ 丢弃标题前噪音: {content}")
                    continue

                # 规则1b: 紧跟在标题后面的 UNKNOWN -> 视为表头噪音 (如 "名专业...") -> 丢弃
                if prev_type == self.LINE_TYPE_TITLE:
                    print(f"    🗑️ 丢弃标题后噪音: {content}")
                    continue
                
                # 规则2: 如果前面有数据行 -> 视为备注 -> 合并
                if last_data_row_idx:
                    current_val = ws.cell(row=last_data_row_idx, column=3).value or ""
                    # 防止无限追加的安全限制
                    if len(current_val) < 800:
                        ws.cell(row=last_data_row_idx, column=3).value = f"{current_val}{content}"
                    continue
                
                # 规则3: 既不在标题后，也没数据可依附 -> 视为正文/段落文字 -> 缓存待写
                pending_text_lines.append(content)
                continue

            # D. 数据行 (GROUP / MAJOR) 处理
            # 只要是数据行，就要先把之前缓存的"碎文字"写出去(如果有的话)
            if l_type in [self.LINE_TYPE_GROUP, self.LINE_TYPE_MAJOR]:
                flush_pending_text()
                
                # 确保表头存在
                if not headers_inserted:
                    if current_row > 1: current_row += 1
                    self._insert_headers(ws, current_row, header_fill, header_font, align_center, border)
                    current_row += 1
                    headers_inserted = True

                # 定义一个内部函数快速写入
                def process_and_write(code, name, raw_text):
                    nonlocal current_row, last_data_row_idx
                    # 使用新的流式解析器
                    entries = self.parse_detailed_line(code, name, raw_text)
                    for e_code, e_name, e_remark in entries:
                        print(f"DEBUG_WRITE: Code='{e_code}', Name='{e_name}'")
                        write_data_row(ws, current_row, [e_code, e_name, e_remark], styles)
                        last_data_row_idx = current_row
                        current_row += 1

                # 若是 GROUP
                if l_type == self.LINE_TYPE_GROUP:
                    for match in self.GROUP_PATTERN.finditer(content):
                        name = match.group(2)
                        name = re.sub(r'(\D)(\d+)', r'\1 \2', name) 
                        process_and_write(match.group(1), name, match.group(3))
                
                # 若是 MAJOR
                elif l_type == self.LINE_TYPE_MAJOR:
                    for match in self.MAJOR_PATTERN.finditer(content):
                        process_and_write(match.group(1), match.group(2), match.group(3))

        flush_pending_text()
        autofit_columns(ws, 3)
        setup_columns(ws)
        return wb

    def parse_detailed_line(self, first_code, first_name, first_remark_start_text):
        """
        基于括号深度的流式解析 (Bracket-Aware Stream Parsing)。
        从 first_remark_start_text 开始扫描，提取当前条目的备注，并检测是否有后续条目。
        
        Args:
            first_code: 第一个条目的代码
            first_name: 第一个条目的名称
            first_remark_start_text: 第一个条目"人"字之后的所有文本（包括"人"字本身如果之前没切分干净的话，但根据正则逻辑，这里传入的应该是 \d+人... 这一段）
        
        Returns:
            [(code, name, remark), ...]
        """
        results = []
        
        # 构造完整的待扫描字符串：为了统一逻辑，我们将第一个条目的 "X人..." 作为流的起点
        # 但要注意，调用方传入的 first_remark_start_text 实际上就是 "matched_group_3"，即 "\d+人..."
        
        # 初始状态
        current_code = first_code
        current_name = first_name
        
        # 待扫描的文本流
        stream = first_remark_start_text
        
        cursor = 0
        length = len(stream)
        bracket_depth = 0
        
        # 当前条目的备注 buffer
        current_remark_buffer = []
        
        while cursor < length:
            char = stream[cursor]
            
            # --- 1. 括号深度维护 ---
            if char in '（(':
                bracket_depth += 1
                # print(f"DEBUG: Char '{char}' at {cursor}. Depth UP -> {bracket_depth}")
            elif char in '）)':
                if bracket_depth > 0:
                    bracket_depth -= 1
                # print(f"DEBUG: Char '{char}' at {cursor}. Depth DOWN -> {bracket_depth}")
            
            # --- 2. 新条目检测 (仅当不在括号内时) ---
            if bracket_depth == 0:
                remaining = stream[cursor:]
                
                # 正则预读 (Lookahead): 
                new_entry_pattern = re.compile(r"^\s*([A-Za-z0-9]{2})?\s*([\u4e00-\u9fa5]{2,}[^\s人]*)\s*(?=\d+\s*人)")
                
                match = new_entry_pattern.match(remaining)
                if match:
                    print(f"    ✂️  Split Entry Found at char '{char}': Code='{match.group(1)}', Name='{match.group(2)}'")
                    # 找到了新条目！
                    # 保存当前条目
                    results.append((current_code, current_name, "".join(current_remark_buffer).strip()))
                    
                    # 更新状态
                    current_code = match.group(1)
                    current_name = match.group(2)
                    current_remark_buffer = [] # 清空 buffer
                    
                    # 移动游标：跳过代码和名字，直接指下一条目的 "人数" 起始处 (因为 new_entry_pattern 没有消耗人数)
                    # match.end() 是匹配到的 Code+Name 的结束位置
                    cursor += match.end()
                    continue
            
            # --- 3. 普通字符处理 ---
            current_remark_buffer.append(char)
            cursor += 1
            
        # 循环结束，保存最后一个条目
        results.append((current_code, current_name, "".join(current_remark_buffer).strip()))
        
        return results

    def _insert_headers(self, ws, row, fill, font, align, border):
        headers = ["专业组/专业代码", "专业组/专业描述", "招生人数/备注"]
        for col, val in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = font
            cell.alignment = align
            cell.border = border

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="安徽省招生计划解析器")
    parser.add_argument("image_path", nargs='?', default=None, help="图片路径")
    parser.add_argument("--output-dir", default="output/anhui", help="输出目录")
    parser.add_argument("--md-input", help="直接解析已有的 Markdown 文件")
    parser.add_argument("--province", help="省份代码（兼容性参数）")
    parser.add_argument("--model", choices=['auto', 'ocr', 'v3'], default='auto',
                        help="OCR 模型选择: auto=自动检测, ocr=强制PaddleOCR, v3=强制StructureV3")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    processor = AnhuiProcessor()
    
    # 设置模型偏好
    processor.model_preference = args.model
    
    if args.md_input:
        print(f"📄 正在从本地 Markdown 验证解析: {args.md_input}")
        with open(args.md_input, 'r', encoding='utf-8') as f:
            md_text = f.read()
        
        base_name = Path(args.md_input).stem
        timestamp = time.strftime("%Y%m%d_%H%M%S")
        
        wb = processor.parse_to_excel(md_text)
        excel_path = Path(args.output_dir) / f"{base_name}_v3_{timestamp}.xlsx"
        wb.save(excel_path)
        print(f"   ✓ Excel 已生成 (本地验证): {excel_path}")
    elif args.image_path:
        processor.run(args.image_path, args.output_dir)
    else:
        print("❌ 请提供图片路径或 --md-input 路径")
